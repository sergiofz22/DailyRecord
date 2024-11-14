
from pathlib import Path
import re
from tkinter import *
import customtkinter
import tkinter as tk
from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage, Label, ttk, messagebox
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

# Definir datos_filas como una variable global
datos_filas = []

OUTPUT_PATH = Path(__file__).parent
ASSETS_PATH = OUTPUT_PATH / Path(r"path")

def calcular_tiempo_esperado(informesRealizados):
    tiempo_esperado = 0
    informes = informesRealizados.split()
    #Definición de tiempos objetivo:
    objetivo_por_tipo={
        'a':1,
        'b':3,
        'c':6
        }
    
    for grupo in informes:
        if grupo.endswith('an'):
            valor_numerico = float(grupo[:-2]) if grupo[:-2].replace('.', '').isdigit() else 1.0
            letra = 'an'
        else:
            valor_numerico = float(grupo[:-1]) if grupo[:-1].replace('.', '').isdigit() else 1.0
            letra = grupo[-1]
        
        tiempo_esperado += valor_numerico * objetivo_por_tipo[letra]
    
    return tiempo_esperado

def sumar_primeros_numeros(cadena):
    # Encontrar todas las secuencias que pueden comenzar con un número (incluyendo decimales) y contienen letras después
    grupos = re.findall(r'\d*\.?\d*[a-zA-Z]+', cadena)
    
    numeros = []
    for grupo in grupos:
        # Si el grupo no comienza con un número, consideramos que el número es 1
        if re.match(r'^[a-zA-Z]', grupo):
            numeros.append(1)
        else:
            # Extraer el número del grupo
            match = re.match(r'\d*\.?\d*', grupo)
            if match:
                numero = match.group()
                # Convertir a float
                if numero:
                    numeros.append(float(numero))
    
    # Sumar los números
    suma = sum(numeros)
    
    return suma

def guardar_datos():
    global datos_filas  # Declarar que estamos usando la variable global

    if not validar_datos():
        return
    
    # Obtener las horas de cada entrada
    horas_totales = 0
    entradas_horas = [entry_informes, entry_ocupacion1, entry_ocupacion2, entry_ocupacion3, entry_otros]
    
    try:
        horas_totales = sum(float(entry.get()) for entry in entradas_horas if entry.get())
    except ValueError:
        tk.messagebox.showerror("Error", "Los campos de horas deben contener números válidos.")
        return

    # Confirmar si las horas no suman 8
    if horas_totales != 8:
        respuesta = tk.messagebox.askyesno(
            "Confirmación", 
            f"La suma de horas totales es {horas_totales}, pero no suman 8 horas. ¿Quieres enviar los datos de todos modos?"
        )
        if not respuesta:
            return  # Cancelar el guardado si la respuesta es "No"

    informesRealizados = entry_informes2.get()
    if informesRealizados:
        tiempo_esperado = calcular_tiempo_esperado(informesRealizados)
        total_informes = sumar_primeros_numeros(informesRealizados)
    else:
        tiempo_esperado = 0
        total_informes = 0

    tiempo_real = entry_informes.get()
    if tiempo_real:
        tiempo_real = float(tiempo_real)
        tiempo_esperado = float(tiempo_esperado)
        eficiencia = (tiempo_esperado / tiempo_real) * 100
    else:
        eficiencia = ""

    # Obtener los valores de iniciales y fecha
    iniciales = entry_iniciales.get()  # Obtener las iniciales del campo correspondiente
    fecha = entry_fecha.get()  # Obtener la fecha del campo correspondiente

    datos_filas = [
        {
            "Iniciales": iniciales,
            "Fecha": fecha,
            "Tipo": "Informes",
            "Detalle": entry_informes.get(),
            "Informes2": entry_informes2.get(),
            "TiempoEsperado": tiempo_esperado,
            "Eficiencia": eficiencia,
            "Observaciones": entry_informes3.get(),
            "Numero de informes": total_informes,
        },
        {
            "Iniciales": iniciales,
            "Fecha": fecha,
            "Tipo": "ocupacion1",
            "Detalle": entry_ocupacion1.get(),
            "Informes2": "",
            "TiempoEsperado": "",
            "Eficiencia": "",
            "Observaciones": entry_ocupacion13.get(),
            "Numero de informes": "",
        },
        {
            "Iniciales": iniciales,
            "Fecha": fecha,
            "Tipo": "Ocupacion2",
            "Detalle": entry_ocupacion2.get(),
            "Informes2": "",
            "TiempoEsperado": "",
            "Eficiencia": "",
            "Observaciones": entry_ocupacion23.get(),
            "Numero de informes": "",
        },
        {
            "Iniciales": iniciales,
            "Fecha": fecha,
            "Tipo": "ocupacion3",
            "Detalle": entry_ocupacion3.get(),
            "Informes2": "",
            "TiempoEsperado": "",
            "Eficiencia": "",
            "Observaciones": entry_ocupacion33.get(),
            "Numero de informes": "",
        },
        {
            "Iniciales": iniciales,
            "Fecha": fecha,
            "Tipo": "Otros",
            "Detalle": entry_otros.get(),
            "Informes2": "",
            "TiempoEsperado": "",
            "Eficiencia": "",
            "Observaciones": entry_otros3.get(),
            "Numero de informes": "",
        }
    ]

    datos_filas_filtradas = [fila for fila in datos_filas if fila["Detalle"] and fila["Detalle"] != "0"]

    if not datos_filas_filtradas:
        tk.messagebox.showinfo("Información", "No hay datos para guardar.")
        return
    
    df = pd.DataFrame(datos_filas_filtradas)

    archivo_excel = Path('').resolve()

    book = load_workbook(archivo_excel)
    sheet = book['Hoja1']

    rows = dataframe_to_rows(df, index=False, header=False)
    for r_idx, row in enumerate(rows, start=sheet.max_row + 1):
        for c_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx)
            if c_idx == 2:
                date_obj = datetime.strptime(value, '%d/%m/%Y')
                cell.value = date_obj
                cell.number_format = 'DD/MM/YYYY'
            elif c_idx == 4:
                try:
                    numeric_value = float(value)
                    cell.value = numeric_value
                    cell.number_format = '0.00'
                except ValueError:
                    cell.value = value
                    cell.number_format = '0.00'
            else:
                cell.value = value
    book.save(archivo_excel)

    if eficiencia:
        tk.messagebox.showinfo("¡GRACIAS!", f"¡Hoy has tenido una eficiencia del {round(eficiencia)}%!.\nEl registro se ha guardado con éxito.")
    else:
        tk.messagebox.showinfo("¡GRACIAS!", "El registro se ha guardado con éxito.")
    
    cancelar()

def validar_datos():
    for entry in [entry_informes, entry_ocupación1, entry_ocupación2, entry_ocupación3, entry_otros]:
        try:
            if entry.get() != "":
                float(entry.get())
        except ValueError:
            tk.messagebox.showerror("Error", "Hay texto en la columna de tiempos.")
            return False

    if not all(re.match(r"^\d*\.?\d+(?:[abc]|an)$", grupo) or grupo in ["a", "b", "c", "an"] for grupo in entry_informes2.get().split()):
        tk.messagebox.showerror("Error", "El campo de tipo de informes debe contener solo a, b, c, an (o múltiplos de ellas). Deben estar separadas por un espacio. Para decimales se usa punto (.) por ejemplo (0.5b 2c b)")
        return False

    if entry_iniciales.get() == "None":
        tk.messagebox.showerror("Error", "El campo 'Iniciales' debe estar cubierto.")
        return False
    
    if not re.match(r"^\d{2}/\d{2}/\d{4}$", entry_fecha.get()):
        tk.messagebox.showerror("Error", "El formato de la fecha debe ser DD/MM/AAAA.")
        return False
    try:
        fecha=datetime.strptime(entry_fecha.get(), "%d/%m/%Y")
    except ValueError:
        tk.messagebox.showerror("Error", "Esa fecha no existe.")
        return False
    if fecha > datetime.today():
        tk.messagebox.showerror("Error", "No puedes introducir datos futuros.")
        return False
    
    return True
    
def cancelar():
    window.destroy()

def borrar_campos():
    entry_iniciales.set("None")
    entry_fecha.delete(0, tk.END)
    entry_informes.delete(0, tk.END)
    entry_informes2.delete(0, tk.END)
    entry_informes3.delete(0, tk.END)
    entry_ocupacion1.delete(0, tk.END)
    entry_ocupacion13.delete(0, tk.END)
    entry_ocupacion2.delete(0, tk.END)
    entry_ocupacion23.delete(0, tk.END)
    entry_ocupacion3.delete(0, tk.END)
    entry_ocupacion33.delete(0, tk.END)
    entry_otros.delete(0, tk.END)
    entry_otros3.delete(0, tk.END)

def ruta_absoluta(relative_path):
    #---------- Descomentar antes de convertir en ejecutable ----------#
    #try:
    #    base_path = sys._MEIPASS
    #except Exception:
    #    base_path = os.path.abspath(".")
    #return os.path.join(base_path, relative_path)
    #---------- Comentar antes de convertir en ejecutable ----------#
    directorio_actual = os.path.dirname(__file__)
    return os.path.join(directorio_actual, relative_path)

def relative_to_assets(path: str) -> Path:
    if getattr(sys, 'frozen', False):
                # Si el script está empaquetado por PyInsocupacion1
        base_path = Path(sys._MEIPASS)  # Ruta base proporcionada por PyInsocupacion1
    else:
        # Si el script se está ejecutando directamente
        base_path = ASSETS_PATH  # Ruta local de los recursos
    return base_path / Path(path)

def load_image(file_path):
                image = Image.open(file_path)
                photo_image=ImageTk.PhotoImage(image)
                return photo_image
            
def main():
    global window, entry_iniciales, entry_fecha, entry_informes, entry_informes2, entry_informes3, entry_ocupacion1, entry_ocupacion13, entry_ocupacion2, entry_ocupacion23, entry_ocupacion3, entry_ocupacion33, entry_otros, entry_otros3

    # Establecer rutas a los recursos
    ruta_icono = relative_to_assets("images/image.ico")
    ruta_boton_1 = relative_to_assets("frame0/button_1.png")
    ruta_boton_2 = relative_to_assets("frame0/button_2.png")
    ruta_boton_3 = relative_to_assets("frame0/button_3.png")
    ruta_imagen_1 = relative_to_assets("images/image2.png")

    #Ventana principal
    window = customtkinter.CTk()
    window.title("Titulo")
    window.iconbitmap(ruta_icono)
    window.geometry("500x500")
    window.resizable(width=False, height=False)
    window.configure(bg = "#FFFF")
    fondo_blanco = tk.Label(window, bg="white")
    fondo_blanco.place(relwidth=1, relheight=1)
    window.grid_columnconfigure(0, weight=1)
    window.grid_rowconfigure(2, weight=1)


    lista_empleados = ['E1', 'Empleado2', 'Empleado3', 'Empleado4', 'Empleado5', 'Empleado6'] 

    #Cargar imagen

    image = Image.open(ruta_imagen_1)
    image = image.resize( (205,82), Image.LANCZOS)

    img = ImageTk.PhotoImage(image)
    lbl_img = Label(window, 
                    image = img,
                    background=  "white")
    lbl_img.place(x=250, y=100)


# Frame 0: Título
    frame0 = tk.Frame(
        window, 
        bg="#0E2E5E",  
        height=85,
        )
    frame0.grid(
        row=0, 
        column=0, 
        columnspan=1, 
        sticky="ew")
    
    title_text = tk.Label(
        frame0, 
        text="Registro diario", 
        fg="#FFFFFF", 
        bg="#0E2E5E", 
        font=("JockeyOne Regular", 24))
    
    title_text.pack(anchor='w', padx=10, pady=25) 

# Frame 1: Iniciales y Fecha
    frame1 = tk.Frame(
        window, 
        bg="white",  
        height=400)
    frame1.grid(
        row=1, 
        column=0, 
        sticky="w")

    label_iniciales = tk.Label(frame1, text="Iniciales:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_iniciales.grid(row=0, column=0, sticky="w", padx=10, pady=10)

    combobox_var = tk.StringVar(value="None")  # set initial value

    def combobox_callback(choice):
        print("combobox dropdown clicked:", choice)

    entry_iniciales = customtkinter.CTkComboBox(frame1,
                                          height=23,
                                          width=93,
                                          font=("KleeOne Regular", 13 * -1),
                                          corner_radius=10,
                                          border_color= "#0E2E5E",
                                          bg_color= "white",
                                          fg_color= "#D9D9D9",
                                          button_color= "#9E9E9E",
                                          dropdown_fg_color= "#0E2E5E",
                                          dropdown_hover_color= "#3E3E3E",
                                          text_color= "#0E2E5E",
                                          dropdown_text_color= "white",
                                          values=['Empleado1', 'Empleado2', 'Empleado3', 'Empleado4', 'Empleado5', 'Empleado6'],
                                          command=combobox_callback,
                                          variable=combobox_var
                                          )
    entry_iniciales.grid(
        row=0, 
        column=1, 
        sticky="w", 
        padx=10, 
        pady=10)

    label_fecha = tk.Label(
        frame1, 
        text="Fecha:", 
        fg="#0E2E5E", 
        bg="white", 
        font=("KleeOne Regular", 13))
    
    label_fecha.grid(
        row=1, 
        column=0, 
        sticky="w", 
        padx=10, 
        pady=10)

    fecha_hoy = datetime.today().strftime('%d/%m/%Y')
    entry_fecha = customtkinter.CTkEntry(
        frame1,
        height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E"
        )
    entry_fecha.grid(row=1, column=1, sticky="w", padx=10, pady=10)
    entry_fecha.insert(0, fecha_hoy)

# Frame 2: Resto de widgets
    frame2 = tk.Frame(window, bg="white", width=700)
    frame2.grid(row=2, column=0, sticky="nsew", padx=10, pady=(20, 10))


    # titulo - TIEMPO [H]
    label_tiempo = tk.Label(frame2, text="Tiempo [h]:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_tiempo.grid(row=0, column=1, sticky="w", padx=10, pady=5)

    # titulo - Tipo de informes
    label_tipo_informes = tk.Label(frame2, text="Tipo informes:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_tipo_informes.grid(row=0, column=2, sticky="w", padx=10, pady=5)

    # titulo - Observaciones
    label_observaciones = tk.Label(frame2, text="Observaciones:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_observaciones.grid(row=0, column=3, sticky="w", padx=10, pady=5)

    # Informes
    label_informes = tk.Label(frame2, text="Informes:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_informes.grid(row=3, column=0, sticky="w", padx=10, pady=5)

        #informes-tiempo
    entry_informes = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_informes.grid(row=3, column=1, padx=10, pady=5)
    entry_informes.get()
        #informes-tipodeinformes
    entry_informes2 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_informes2.grid(row=3, column=2, padx=10, pady=5)
    entry_informes2.get()
        #informes-observaciones
    entry_informes3 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_informes3.grid(row=3, column=3, padx=10, pady=5)
    entry_informes3.get()

    # ocupacion1
    label_ocupacion1 = tk.Label(frame2, text="ocupacion1:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_ocupacion1.grid(row=4, column=0, sticky="w", padx=10, pady=5)

        #ocupacion1-tiempo
    entry_ocupacion1 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_ocupacion1.grid(row=4, column=1, padx=10, pady=5)
    entry_ocupacion1.get()
        #ocupacion1-observaciones
    entry_ocupacion13 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_ocupacion13.grid(row=4, column=3, padx=10, pady=5)
    entry_ocupacion13.get()

    # Ocupacion2
    label_ocupacion2 = tk.Label(frame2, text="Ocupacion2:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_ocupacion2.grid(row=5, column=0, sticky="w", padx=10, pady=5)
        #ocupacion2-tiempo
    entry_ocupacion2 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_ocupacion2.grid(row=5, column=1, padx=10, pady=5)
    entry_ocupacion2.get()
        #ocupacion2-observaciones
    entry_ocupacion23 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_ocupacion23.grid(row=5, column=3, padx=10, pady=5)
    entry_ocupacion23.get()

    # ocupacion3
    label_ocupacion3 = tk.Label(frame2, text="ocupacion3:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_ocupacion3.grid(row=6, column=0, sticky="w", padx=10, pady=5)
        #ocupacion3-tiempo
    entry_ocupacion3 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_ocupacion3.grid(row=6, column=1, padx=10, pady=5)
    entry_ocupacion3.get()
        #ocupacion3-observaciones
    entry_ocupacion33 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_ocupacion33.grid(row=6, column=3, padx=10, pady=5)
    entry_ocupacion33.get()

    # Otros
    label_otros = tk.Label(frame2, text="Otros:", fg="#0E2E5E", bg="white", font=("KleeOne Regular", 13))
    label_otros.grid(row=10, column=0, sticky="w", padx=10, pady=5)
        #otros-tiempo
    entry_otros = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_otros.grid(row=10, column=1, padx=10, pady=5)
    entry_otros.get()
        #otros-observaciones
    entry_otros3 = customtkinter.CTkEntry(frame2, height=23, width=93, font=("KleeOne Regular", 13), corner_radius=10, border_color="#0E2E5E", bg_color="white", fg_color="#D9D9D9", text_color="#0E2E5E")
    entry_otros3.grid(row=10, column=3, padx=10, pady=5)
    entry_otros3.get()

# Frame 3: Botones 

    frame3 = tk.Frame(window, bg="white", height=65)
    frame3.grid(row=3, column=0, sticky="ew", padx=10, pady=(10, 10))

    button_image_1 = PhotoImage(file=ruta_boton_1)
    btn_cancelar = Button(frame3, image=button_image_1, borderwidth=0, highlightthickness=0, command=cancelar, relief="flat")
    btn_cancelar.pack(side="left", padx=26, pady=10)

    button_image_2 = PhotoImage(file=ruta_boton_2)
    btn_borrar = Button(frame3, image=button_image_2, borderwidth=0, highlightthickness=0, command=borrar_campos, relief="flat")
    btn_borrar.pack(side="left", padx=26,  pady=10)

    button_image_3 = PhotoImage(file=ruta_boton_3)
    btn_guardar = Button(frame3, image=button_image_3, borderwidth=0, highlightthickness=0, command=guardar_datos, relief="flat")
    btn_guardar.pack(side="left", padx=25,  pady=10)

    window.resizable(False, False)
    window.mainloop()

if __name__ == "__main__":
    main()
